#!/usr/bin/env python
# -*- coding:utf-8 -*-
# _AUTHOR_  : zhujingxiu
# _DATE_    : 2018/12/08
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


class read4Linq(object):

    def __init__(self, xls_dir, sheet_keyword, output=False):
        self.xls_dir = xls_dir
        self.sheet_keyword = sheet_keyword
        self.output = output if output else self.xls_dir

    def readDir(self):
        num = 1
        detail_rows = []
        for _file in os.listdir(self.xls_dir):
            ret = self.readFile(_file)
            detail_rows.append({'sn': num, 'items': ret, 'length': len(ret)})
            num += 1
        return detail_rows

    def checkingDir(self):
        detail_rows = []
        for _file in os.listdir(self.xls_dir):
            ret = self.readFile(_file, True)
            if ret:
                detail_rows.extend(ret)
        return detail_rows

    def readFile(self, filename, checking=False):
        wb = load_workbook(os.path.join(self.xls_dir, filename), data_only=True)
        firstsheet = wb.worksheets[0]
        rows = []
        _store = firstsheet["A2"].value
        _c_cell = firstsheet["C1"].value
        if checking:
            error = []
            if not _store:
                error.append("没有店铺名")
            if not _c_cell:
                error.append("第三栏没有命名")
            if error:
                rows.append({
                    'store': filename,
                    'error': " , ".join(error)
                })
        else:
            _principal_total = _fee_total = 0

            for item in firstsheet["B"]:
                _buyer = item.value

                if _c_cell == '电话':
                    _principal = firstsheet["D%s" % item.row].value
                    _fee = firstsheet["E%s" % item.row].value
                else:
                    _principal = firstsheet["C%s" % item.row].value
                    _fee = firstsheet["D%s" % item.row].value

                if _buyer:
                    rows.append({
                        'store': _store,
                        'buyer': item.value,
                        'principal': _principal,
                        'fee': _fee,
                    })
                else:
                    _principal_total = _principal
                    _fee_total = _fee

            for i in range(3):
                rows.append({
                    'store': '',
                    'buyer': '',
                    'principal': 0,
                    'fee': 0,
                })
            rows.append({
                'store': '',
                'buyer': '合计',
                'principal': _principal_total,
                'fee': _fee_total,
            })
            for i in range(2):
                rows.append({
                    'store': '',
                    'buyer': '',
                    'principal': 0,
                    'fee': 0,
                })
            rows.pop(0)
        return rows

    def genSheets(self, result):
        from openpyxl.styles import Font, colors, Alignment
        wb = Workbook()
        basesheet = wb['Sheet']
        basesheet.title = r'%s汇总' % self.sheet_keyword

        detailsheet = wb.create_sheet(r'%s刷单明细' % self.sheet_keyword, index=1)

        detailsheet["A1"] = "序号"
        detailsheet["B1"] = "店铺名称"
        detailsheet["C1"] = "买手号"
        detailsheet["D1"] = "任务本金"
        detailsheet["E1"] = "任务佣金"
        detailsheet["F1"] = "平台佣金"
        detailsheet["G1"] = "毛利"
        detailsheet["H1"] = "笔数"
        bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
        detailsheet.row_dimensions[1].height = 40
        detailsheet.row_dimensions[1].font = bold_itatic_24_font
        row_num = 2
        all_principal = all_fee = all_profit = all_value = all_buyer = 0
        for item in result:
            sn = item['sn']
            _first_num = 0
            _buyer_num = 0
            _sn_profit = 0
            _sn_value = 0
            for row in item['items']:
                if row['buyer'] == '合计':
                    detailsheet["A%s" % str(row_num)] = ''
                    detailsheet["B%s" % str(row_num)] = row['store']
                    detailsheet["C%s" % str(row_num)] = row['buyer']
                    detailsheet["D%s" % str(row_num)] = row['principal']
                    detailsheet["E%s" % str(row_num)] = row['fee']
                    detailsheet["F%s" % str(row_num)] = _sn_value
                    detailsheet["G%s" % str(row_num)] = _sn_profit
                elif row['store'] and row['buyer'] and row['principal']:
                    _first_num = row_num if not _first_num else _first_num
                    try:
                        _value = row['principal'] * 0.01 + 12
                        _profit = float(row['fee']) - _value
                        detailsheet["A%s" % str(row_num)] = sn
                        detailsheet["B%s" % str(row_num)] = row['store']
                        detailsheet["C%s" % str(row_num)] = row['buyer']
                        detailsheet["D%s" % str(row_num)] = row['principal']
                        detailsheet["E%s" % str(row_num)] = row['fee']
                        detailsheet["F%s" % str(row_num)] = _value
                        detailsheet["G%s" % str(row_num)] = _profit
                        _buyer_num += 1
                        all_principal += float(row['principal'])
                        all_fee += float(row['fee'])
                        all_value += _value
                        all_profit += _profit
                        all_buyer += 1
                        _sn_profit += _profit
                        _sn_value += _value
                    except Exception as e:
                        print(sn, row['store'], e)
                row_num += 1
            if _first_num:
                detailsheet.merge_cells('H%s:H%s' % (str(_first_num), str(_first_num + _buyer_num - 1)))
                detailsheet["H%s" % str(_first_num)] = _buyer_num
                detailsheet["H%s" % str(_first_num)].alignment = Alignment(horizontal='center', vertical='center')
        detailsheet.column_dimensions['B'].width = 20
        detailsheet.column_dimensions['C'].width = 30

        row_num += 1

        detailsheet["A%s" % str(row_num)] = '汇总'
        detailsheet["B%s" % str(row_num)] = '汇总'
        detailsheet["C%s" % str(row_num)] = '汇总'
        detailsheet["D%s" % str(row_num)] = all_principal
        detailsheet["E%s" % str(row_num)] = all_fee
        detailsheet["F%s" % str(row_num)] = all_profit
        detailsheet["G%s" % str(row_num)] = all_value
        detailsheet["H%s" % str(row_num)] = all_buyer

        detailsheet.merge_cells('A%s:C%s' % (str(row_num), str(row_num)))
        detailsheet["A%s" % str(row_num)].alignment = Alignment(horizontal='center', vertical='center')
        xls_file = r'%s/%s-%s汇总.xlsx' % (self.output, self.sheet_keyword, datetime.now().strftime('%Y.%m.%d%H%M%S'))
        wb.save(xls_file)
        return xls_file
